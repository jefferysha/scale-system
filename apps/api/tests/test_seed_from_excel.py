"""seed-from-excel 单元测试（用临时小 Excel + testcontainers PG）."""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest
from sqlalchemy import select

# 让脚本能 import：tests/test_*.py → apps/api/tests → repo root → scripts/
REPO_ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(REPO_ROOT / "scripts"))

from seed_from_excel.cups_loader import load_cups  # noqa: E402
from seed_from_excel.projects_loader import load_projects  # noqa: E402
from seed_from_excel.records_loader import load_records  # noqa: E402

from scale_api.models.cup import Cup  # noqa: E402
from scale_api.models.cup_calibration import CupCalibration  # noqa: E402
from scale_api.models.project import Project  # noqa: E402
from scale_api.models.record import WeighingRecord  # noqa: E402
from scale_api.models.vertical import Vertical  # noqa: E402


@pytest.fixture
def tiny_excel(tmp_path: Path) -> Path:
    """构造一个简化但覆盖关键场景的 Excel。

    覆盖：
    - 项目库 1 行（空标题列匹配真实 Excel 头）
    - 杯库 3 行（1 条无当前杯重 → 失败计数；2 条正常）
    - 一个老 sheet（用 bh1/bs1 typo，无容积列）
    - 一个新 sheet（用 bh10/bs10，有容积列）
    """
    wb = openpyxl.Workbook()
    # 项目库
    ws_proj = wb.active
    ws_proj.title = "项目库"
    ws_proj.append([None, "项目", "建立日期", "备注"])
    ws_proj.append([1, "TestProj-old", "2007年12月24日", "测试项目老格式"])
    ws_proj.append([2, "TestProj-new", datetime(2016, 12, 13), None])

    # 杯库
    ws_cup = wb.create_sheet("杯库")
    ws_cup.append(["杯号", "当前杯重", "最新率定日期杯重", "上次杯重", "上次率定日期"])
    ws_cup.append(["TC-001", 35.2480, datetime(2025, 8, 1), None, None])
    ws_cup.append(
        ["TC-002", 41.6712, datetime(2024, 5, 15), 41.6713, datetime(2023, 3, 10)]
    )
    ws_cup.append(["TC-BAD", None, datetime(2024, 5, 15), None, None])  # 失败行

    # 老格式 sheet（bh1/bs1 typo + 无容积列），sheet 名要在 PROJECT_SHEETS 列表里
    ws_old = wb.create_sheet("TestProj-old")
    ws_old.append(
        [
            "垂线号", "潮型", "日期", "水深",
            "c0", "c2", "c4", "c6", "c8", "c10",
            "bh0", "bh2", "bh4", "bh6", "bh8", "bh1",
            "bs0", "bs2", "bs4", "bs6", "bs8", "bs1",
            "开始时间", "结束时间",
        ]
    )
    ws_old.append(
        [
            "V-1", "大潮", datetime(2016, 11, 13, 1, 0), 10,
            0.30, 0.31, 0.32, 0.33, 0.34, 0.35,
            "TC-001", "TC-002", "TC-001", "TC-002", "TC-001", "TC-002",
            45.0, 46.0, 47.0, 48.0, 49.0, 50.0,
            None, None,
        ]
    )

    # 新格式 sheet（bh10/bs10 + 有容积列）
    ws_new = wb.create_sheet("TestProj-new")
    ws_new.append(
        [
            "垂线号", "潮型", "日期", "水深",
            "c0", "c2", "c4", "c6", "c8", "c10",
            "bh0", "bh2", "bh4", "bh6", "bh8", "bh10",
            "bs0", "bs2", "bs4", "bs6", "bs8", "bs10",
            "开始时间", "结束时间", "容积",
        ]
    )
    ws_new.append(
        [
            "V-2", "小潮", datetime(2016, 11, 14, 2, 0), 12,
            0.10, 0.11, 0.12, 0.13, 0.14, 0.15,
            "TC-002", "TC-001", "TC-002", "TC-001", "TC-002", "TC-001",
            41.0, 42.0, 43.0, 44.0, 45.0, 46.0,
            None, None, 500,
        ]
    )

    f = tmp_path / "tiny.xlsx"
    wb.save(f)
    return f


@pytest.fixture
def tiny_wb(tiny_excel: Path):
    return openpyxl.load_workbook(tiny_excel, data_only=True, read_only=True)


@pytest.mark.asyncio
async def test_load_projects_inserts_unique(session, tiny_wb) -> None:
    ok, fail = await load_projects(session, tiny_wb, dry_run=False)
    await session.commit()
    assert ok == 2
    assert fail == 0
    rows = (await session.scalars(select(Project))).all()
    assert {p.name for p in rows} == {"TestProj-old", "TestProj-new"}

    # 重跑：幂等（已存在算成功，且不重复插）
    ok2, fail2 = await load_projects(session, tiny_wb, dry_run=False)
    await session.commit()
    assert ok2 == 2
    assert fail2 == 0
    rows2 = (await session.scalars(select(Project))).all()
    assert len(rows2) == 2


@pytest.mark.asyncio
async def test_load_cups_with_calibration_and_failure(session, tiny_wb) -> None:
    ok, fail = await load_cups(session, tiny_wb, dry_run=False)
    await session.commit()
    assert ok == 2  # TC-001, TC-002
    assert fail == 1  # TC-BAD（无当前杯重）

    cups = (await session.scalars(select(Cup))).all()
    assert {c.cup_number for c in cups} == {"TC-001", "TC-002"}

    # TC-001 只有 current → 1 条 calibration
    tc001 = next(c for c in cups if c.cup_number == "TC-001")
    cals_001 = (
        await session.scalars(
            select(CupCalibration).where(CupCalibration.cup_id == tc001.id)
        )
    ).all()
    assert len(cals_001) == 1
    assert cals_001[0].method == "import:current"

    # TC-002 有 current + previous → 2 条
    tc002 = next(c for c in cups if c.cup_number == "TC-002")
    cals_002 = (
        await session.scalars(
            select(CupCalibration).where(CupCalibration.cup_id == tc002.id)
        )
    ).all()
    assert len(cals_002) == 2
    assert {c.method for c in cals_002} == {"import:current", "import:previous"}

    # 重跑幂等：cups / cup_calibrations 数量不变
    ok2, _ = await load_cups(session, tiny_wb, dry_run=False)
    await session.commit()
    assert ok2 == 2
    cups_after = (await session.scalars(select(Cup))).all()
    assert len(cups_after) == 2
    cals_after = (await session.scalars(select(CupCalibration))).all()
    assert len(cals_after) == 3


@pytest.mark.asyncio
async def test_load_records_handles_typo_and_volume_default(
    session, tiny_wb, monkeypatch
) -> None:
    """records_loader 应：
    - 兼容 bh1/bs1 typo → 取到 1.0 位置点位
    - 没"容积"列时默认 1000mL
    - 自动建 vertical
    - client_uid 稳定，重跑幂等
    """
    # 把 PROJECT_SHEETS 临时换成 fixture 里的 sheet 名
    import seed_from_excel.records_loader as rl

    monkeypatch.setattr(rl, "PROJECT_SHEETS", ["TestProj-old", "TestProj-new"])

    # 先建 projects + cups 才能挂关系
    await load_projects(session, tiny_wb, dry_run=False)
    await load_cups(session, tiny_wb, dry_run=False)
    await session.commit()

    ok, fail = await load_records(session, tiny_wb, dry_run=False)
    await session.commit()
    assert fail == 0
    assert ok == 2

    records = (await session.scalars(select(WeighingRecord))).all()
    assert len(records) == 2

    # 老 sheet 没"容积"列 → 默认 1000
    old_proj = (
        await session.scalars(select(Project).where(Project.name == "TestProj-old"))
    ).first()
    old_rec = next(r for r in records if r.project_id == old_proj.id)
    assert old_rec.volume_ml == 1000
    # 6 个点位都解出（bh1/bs1 兼容）
    assert len(old_rec.points) == 6
    positions = [p["pos"] for p in old_rec.points]
    assert "1.0" in positions

    # 新 sheet 有"容积"=500
    new_proj = (
        await session.scalars(select(Project).where(Project.name == "TestProj-new"))
    ).first()
    new_rec = next(r for r in records if r.project_id == new_proj.id)
    assert new_rec.volume_ml == 500
    assert len(new_rec.points) == 6

    # 自动建 vertical
    verticals = (await session.scalars(select(Vertical))).all()
    assert {v.code for v in verticals} == {"V-1", "V-2"}

    # 幂等：重跑数字一致
    ok2, fail2 = await load_records(session, tiny_wb, dry_run=False)
    await session.commit()
    assert ok2 == 2
    assert fail2 == 0
    records_after = (await session.scalars(select(WeighingRecord))).all()
    assert len(records_after) == 2
